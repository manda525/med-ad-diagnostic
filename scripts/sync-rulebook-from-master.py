#!/usr/bin/env python3
"""正本 rulebook_master（xlsx）→ rulebook.json（タプル配列）

  python3 scripts/sync-rulebook-from-master.py [--master PATH] [--version vNN] [--note "..."]

これまで正本から rulebook.json への転記に決まった手順が無く、v24 を最後に
5バージョン取り残された（v25 の百貨店106件、v26 の794-795、v27 の796-797、
v28 の798）。手で写す限り同じことが起きるので、正本を読んで機械的に書き出す。

正本は拡張子なしの xlsx で、openpyxl は拡張子で弾く。一時ファイルへ .xlsx として
コピーしてから開く（CLAUDE.md の SOP と同じ回避策）。

■ 百貨店_店頭厳しめ を RB から分ける理由
rule_id 688-793 の106件は三越伊勢丹グループの店頭基準に由来する私的基準で、
法令ではない。百貨店の店頭に卸す案件だけが従う条件付きプロファイルであって、
これを常時適用すると通常の広告がほぼ全滅する。フラグで切り替える形にすると
既定値の取り違え一発で全案件に降ってくるので、そもそも既定の診断セットとは
別の配列（DEP）に出し、build-rulebook-v2.mjs が別ファイルへ書き出す。
既定 OFF が設定でなく構造で保証される。

EX（901-919）と CS（C1-01…）は正本に無い独自ソースなので、既存の
rulebook.json からそのまま引き継ぐ。
"""

import argparse
import json
import os
import re
import shutil
import sys
import tempfile

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl が要ります: pip3 install openpyxl")

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

DEFAULT_MASTER = os.path.expanduser(
    "~/Library/CloudStorage/GoogleDrive-m.and.a2016328@gmail.com/"
    "マイドライブ/00_System_管理/rulebook_master"
)

# 01_ルールブック の列 → タプル [id, ng, risk, genre, comment, ok, law, jcia]
COL = {"id": 1, "ng": 2, "ok": 3, "risk": 4, "genre": 5, "law": 7, "comment": 9, "jcia": 10}

# 条件付きプロファイル。既定の診断セットには入れない。
PROFILE_GENRE = re.compile(r"^百貨店")


def cell(ws, row, name):
    v = ws.cell(row, COL[name]).value
    return "" if v is None else v


def read_master(path):
    with tempfile.TemporaryDirectory() as td:
        tmp = os.path.join(td, "master.xlsx")
        shutil.copyfile(path, tmp)
        wb = openpyxl.load_workbook(tmp, data_only=True)
        meta = {}
        wm = wb["_META"]
        for r in range(1, wm.max_row + 1):
            k = wm.cell(r, 1).value
            if k:
                meta[str(k)] = wm.cell(r, 2).value
        ws = wb["01_ルールブック"]
        rb, dep = [], []
        for r in range(2, ws.max_row + 1):
            rid = cell(ws, r, "id")
            if rid == "":
                continue
            genre = str(cell(ws, r, "genre"))
            row = [
                int(rid),
                str(cell(ws, r, "ng")),
                int(cell(ws, r, "risk") or 0),
                genre,
                str(cell(ws, r, "comment")),
                str(cell(ws, r, "ok")),
                str(cell(ws, r, "law")),
                str(cell(ws, r, "jcia")),
            ]
            (dep if PROFILE_GENRE.match(genre) else rb).append(row)
    return meta, rb, dep


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--master", default=os.environ.get("RULEBOOK_MASTER", DEFAULT_MASTER))
    ap.add_argument("--version", help="rulebook.json 側の版数（例 v19）。省略時は据え置き")
    ap.add_argument("--note", default="", help="meta.note に入れる説明")
    ap.add_argument("--dry-run", action="store_true")
    args = ap.parse_args()

    if not os.path.exists(args.master):
        sys.exit(f"正本が見つかりません: {args.master}")

    meta, rb, dep = read_master(args.master)
    master_ver = meta.get("version")
    master_total = meta.get("total_rules")
    print(f"正本: v{master_ver} / total_rules={master_total} / last_updated={meta.get('last_updated')}")

    if len(rb) + len(dep) != master_total:
        sys.exit(f"件数不一致: 読み取り {len(rb)}+{len(dep)}={len(rb)+len(dep)} vs _META {master_total}")

    dst = os.path.join(ROOT, "rulebook.json")
    cur = json.load(open(dst, encoding="utf-8"))

    ids = [r[0] for r in rb + dep]
    dupes = {i for i in ids if ids.count(i) > 1}
    if dupes:
        sys.exit(f"正本に rule_id 重複: {sorted(dupes)[:10]}")

    short = [(r[0], t) for r in rb for t in re.split(r"[／/・]", r[1]) if 0 < len(t.strip()) <= 2]
    if short:
        print(f"⚠️ 2文字以下の照合語 {len(short)}件（誤爆候補）: {short[:8]}")

    out = {
        "meta": {
            **cur["meta"],
            "version": args.version or cur["meta"].get("version"),
            "previous_version": cur["meta"].get("version"),
            "rule_count": len(rb),
            "profile_depstore_count": len(dep),
            "ex_count": len(cur.get("EX", [])),
            "cs_count": len(cur.get("CS", [])),
            "master_version": master_ver,
            "master_total_rules": master_total,
            "master_last_updated": meta.get("last_updated"),
            "updated": str(meta.get("last_updated") or ""),
            "note": args.note or cur["meta"].get("note", ""),
        },
        "RB": rb,
        "DEP": dep,
        "EX": cur.get("EX", []),
        "CS": cur.get("CS", []),
    }

    print(f"RB {len(cur.get('RB', []))} → {len(rb)} 件 ／ DEP（百貨店・既定OFF） {len(dep)} 件")
    print(f"EX {len(out['EX'])} 件 ／ CS {len(out['CS'])} 件（据え置き）")

    if args.dry_run:
        print("dry-run のため書き込みませんでした")
        return

    with open(dst, "w", encoding="utf-8") as f:
        json.dump(out, f, ensure_ascii=False)
    print(f"書き出し: {dst}")
    print("次に: npm run build:rulebook && npm test")


if __name__ == "__main__":
    main()
